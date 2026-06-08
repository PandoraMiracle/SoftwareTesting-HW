"""Sync Testcases.xlsx with hw1.md Requirement 3 data."""
import openpyxl
from openpyxl.styles import Alignment

PATH = r"c:\DiskD\HCMUS\Semester9\SoftwareTesting\SoftwareTesting-HW\HW1\Testcases.xlsx"

STUDENT_ID = "23127271"
STUDENT_NAME = "Võ Ngọc Bích Trâm"
TEST_DATE = "2026/06/08"
BUILD = "GOOJODOQ GFS006 — Build 1 (physical unit)"

REQUIREMENTS = [
    (1, "R001", "Power & Charging", "Type-C charging, power-on, charge interruption, charge-while-use"),
    (2, "R002", "Speed & Airflow", "100 speed levels, max airflow, smooth adjustment"),
    (3, "R003", "Display & UI", "LED speed display accuracy and retention after power cycle"),
    (4, "R004", "Cooling Plate", "Semiconductor cooling plate temperature reduction"),
    (5, "R005", "Battery & Runtime", "Continuous runtime, low-battery warning/shutdown"),
    (6, "R006", "Noise", "Noise level at low speed vs ≤25 dB spec"),
    (7, "R007", "Ergonomics & Build", "Handheld comfort, desk-mode stability"),
    (8, "R008", "Edge Cases (AI-missed)", "Charge-while-use, charging indicator, drop test"),
]

TEST_TYPES = [
    (1, "Functional", "Functional Testing"),
    (2, "GUI&Usability", "GUI & Usability Testing"),
    (3, "Performance", "Performance Testing"),
]

TEST_CASES = [
    {
        "id": "TC-01",
        "type": "Functional",
        "req": "Power & Charging",
        "objective": "Verify device powers on and produces airflow at default speed",
        "precondition": "Fan fully charged; power slide switch at OFF; Type-C cable disconnected; cold mode Toggle OFF",
        "steps": "1. Slide the power switch from OFF to ON.\n2. Observe the LED display.\n3. Hold hand 10 cm from the air outlet for 5 s.",
        "expected": "Device powers on within 3 s; LED shows a speed level; steady airflow at outlet; no abnormal noise or vibration.",
        "actual": "As expected",
        "status": "Passed",
        "bug": None,
        "executed": True,
        "video": "https://youtube.com/shorts/LB2VY5YWCAI",
        "ai": "Yes",
        "remark": "Video recorded with voice narration",
    },
    {
        "id": "TC-02",
        "type": "Functional",
        "req": "Power & Charging",
        "objective": "Verify Type-C charging from empty/low battery",
        "precondition": "Battery at ≤10% or depleted; standard 5 V/2 A Type-C adapter and cable; power slide switch at OFF",
        "steps": "1. Connect Type-C cable to fan and adapter.\n2. Note start time and charging indicator.\n3. Leave charging until full (~2.5 h).\n4. Disconnect cable and Slide power switch to ON.",
        "expected": "Charging indicator activates within 10 s; full charge in ~2.5 h (±15 min); device powers on normally after disconnect.",
        "actual": "As expected",
        "status": "Passed",
        "bug": None,
        "executed": False,
        "video": None,
        "ai": "Yes",
        "remark": None,
    },
    {
        "id": "TC-03",
        "type": "Functional",
        "req": "Speed & Airflow",
        "objective": "Verify maximum wind speed (level 100) delivers strong airflow",
        "precondition": "Fully charged; handheld or desk mode on stable surface; cold mode Toggle OFF",
        "steps": "1. Slide power switch to ON.\n2. Press speed + button until LED shows 100.\n3. Confirm LED shows 100.\n4. Hold tissue paper or hand 20 cm from outlet for 10 s.",
        "expected": "LED displays 100; airflow stronger than level 50; motor runs smoothly; airflow consistent with 10 m/s / 13,000 RPM spec.",
        "actual": "As expected",
        "status": "Passed",
        "bug": None,
        "executed": True,
        "video": "https://youtube.com/shorts/DTeOIx5WkN4",
        "ai": "Yes",
        "remark": "Video recorded with voice narration",
    },
    {
        "id": "TC-04",
        "type": "Functional",
        "req": "Speed & Airflow",
        "objective": "Verify 100 speed levels adjust smoothly",
        "precondition": "Fully charged; power slide switch at ON; speed at level 1",
        "steps": "1. Press speed + from level 1 to 100, pausing 1 s every level.\n2. Record LED value at each pause.\n3. Press speed − from 100 back to 1.",
        "expected": "Each press changes speed monotonically; LED increments/decrements; motor transitions without stutter or sudden stops.",
        "actual": "As expected",
        "status": "Passed",
        "bug": None,
        "executed": False,
        "video": None,
        "ai": "Yes",
        "remark": None,
    },
    {
        "id": "TC-05",
        "type": "Functional",
        "req": "Display & UI",
        "objective": "Verify LED display accuracy at representative speed levels",
        "precondition": "Fully charged; power slide switch at ON",
        "steps": "1. Press speed − until LED shows 1; note reading.\n2. Press speed + until LED shows 50; note reading.\n3. Press speed + until LED shows 100; note reading.\n4. Slide power OFF then ON; check if last speed level is retained.",
        "expected": "LED shows 1, 40, and 100 respectively with no digit flicker; display remains readable for 30 s at each level; after power cycle, behavior matches user-manual spec (retain or reset to default).",
        "actual": "Last speed level is not retained.",
        "status": "Failed",
        "bug": "B001",
        "executed": True,
        "video": "https://youtube.com/shorts/Ik0BMi2CiCQ",
        "ai": "Yes",
        "remark": "Defect logged — GitHub commit 5a33eae",
    },
    {
        "id": "TC-06",
        "type": "Functional",
        "req": "Cooling Plate",
        "objective": "Verify semiconductor cooling plate lowers surface temperature",
        "precondition": "Fully charged; room temp 25–30 °C; power slide switch at ON; cold mode Toggle OFF",
        "steps": "1. Slide power switch to ON.\n2. Toggle cold mode ON.\n3. Press speed + until LED shows 40–50; run 3 min.\n4. Lightly touch cooling plate for 2 s.\n5. Compare plate temp to plastic housing.",
        "expected": "Cooling plate noticeably cooler than body within 3 min; no sharp heat; fan continues; no indoor condensation.",
        "actual": "As expected",
        "status": "Passed",
        "bug": None,
        "executed": True,
        "video": "https://youtube.com/shorts/w-RW93c0CjI",
        "ai": "Yes",
        "remark": "Video recorded with voice narration",
    },
    {
        "id": "TC-07",
        "type": "Performance",
        "req": "Battery & Runtime",
        "objective": "Verify continuous runtime at medium speed meets 2.5–8 h spec",
        "precondition": "Fully charged (100%); speed at level 50; stopwatch ready; cold mode Toggle OFF",
        "steps": "1. Slide power ON; Press speed +/− until LED shows 50; start stopwatch.\n2. Run continuously without changing settings.\n3. Record time until auto shut-off or speed drop.\n4. Note final battery state.",
        "expected": "Runs ≥2.5 h before shutdown; runtime at level 50 within 2.5–8 h range; graceful shut-off.",
        "actual": "As expected",
        "status": "Passed",
        "bug": None,
        "executed": False,
        "video": None,
        "ai": "Yes",
        "remark": None,
    },
    {
        "id": "TC-08",
        "type": "Performance",
        "req": "Noise",
        "objective": "Verify noise level at low speed is within quiet-use expectation (<25 dB)",
        "precondition": "Fully charged; quiet room (background <30 dB); power slide switch at ON; cold mode Toggle OFF",
        "steps": "1. Place fan on desk; Slide power switch to ON.\n2. Press speed − until LED shows 1.\n3. Measure sound at 1 m using decibel-meter app.\n4. Run 2 min and listen for rattling.",
        "expected": "Noise at level 1 is ≤25 dB (or suitable for office use); no rattling; airflow still perceptible.",
        "actual": "Decibel-meter app measured ~34 dB at 1 m at speed level 1, exceeding ≤25 dB spec; no rattling; airflow perceptible.",
        "status": "Failed",
        "bug": "B004",
        "executed": False,
        "video": None,
        "ai": "Yes",
        "remark": "Defect logged — GitHub commit 5a86e33",
    },
    {
        "id": "TC-09",
        "type": "GUI&Usability",
        "req": "Ergonomics & Build",
        "objective": "Verify handheld ergonomics and thermal comfort during extended use",
        "precondition": "Fully charged; speed at level 60; 15-min test session; cold mode Toggle OFF",
        "steps": "1. Slide power ON; Press speed + until LED shows 60.\n2. Hold fan in handheld mode for 15 min.\n3. Note grip comfort, weight fatigue, and housing temperature.",
        "expected": "Weight (~320 g) comfortable for 15 min; housing not uncomfortably hot; grip secure; airflow easy to aim.",
        "actual": "Weight comfortable for 15 min; housing warm but not hot; grip secure; airflow easy to aim.",
        "status": "Passed",
        "bug": None,
        "executed": False,
        "video": None,
        "ai": "Yes",
        "remark": None,
    },
    {
        "id": "TC-10",
        "type": "GUI&Usability",
        "req": "Ergonomics & Build",
        "objective": "Verify desk/tabletop mode stability at maximum speed on flat surface",
        "precondition": "Fully charged; fan in desk-stand configuration; flat hard tabletop; cold mode Toggle OFF",
        "steps": "1. Place fan in desk mode on level table.\n2. Slide power switch to ON.\n3. Press speed + until LED shows 100.\n4. Observe 3 min for movement or tip-over.\n5. Lightly tap table edge.",
        "expected": "Fan upright 5 min with no walking/sliding; does not tip when table lightly tapped.",
        "actual": "Fan remained upright during normal operation; tipped over after tabletop was lightly tapped.",
        "status": "Failed",
        "bug": "B002",
        "executed": True,
        "video": "https://youtube.com/shorts/lG8OZdGog2k",
        "ai": "Yes",
        "remark": "Defect logged — GitHub commit 263c4f3",
    },
    {
        "id": "TC-11",
        "type": "Functional",
        "req": "Edge Cases (AI-missed)",
        "objective": "Edge case: Verify charge-while-use behavior (simultaneous charging and operation)",
        "precondition": "Battery ~30–50%; 5 V/2 A Type-C adapter; cold mode Toggle OFF; indoor 25–30 °C",
        "steps": "1. Slide power ON; Press speed + until LED shows 50.\n2. Connect Type-C charger while fan is running.\n3. Run 15 min; every 5 min record housing temp, LED, motor sound.\n4. Note whether battery level changes during use-while-charging.",
        "expected": "Safe housing temp; stable motor at level 50; no auto shut-off; charging behavior documented.",
        "actual": "As expected",
        "status": "Passed",
        "bug": None,
        "executed": False,
        "video": None,
        "ai": "No",
        "remark": "Student-added edge case; AI did not generate",
    },
    {
        "id": "TC-12",
        "type": "Functional",
        "req": "Edge Cases (AI-missed)",
        "objective": "Edge case: Verify battery status indicator during charging and at 100% full charge",
        "precondition": "Battery ≤20% or depleted; power OFF; Type-C adapter connected; spec silent on indicator",
        "steps": "1. Connect Type-C with device OFF; note LED/battery indicator.\n2. Observe indicator every 30 min during charging (~2.5 h).\n3. Record behavior at mid-charge and at full charge.\n4. Disconnect; Slide power ON; confirm normal operation.",
        "expected": "Charging state visible; indicator signals full charge at 100%; device operates normally after charge.",
        "actual": "Charging status displayed during charging; no indication when battery reaches full charge.",
        "status": "Failed",
        "bug": "B003",
        "executed": False,
        "video": None,
        "ai": "No",
        "remark": "Student-added edge case; defect — GitHub commit fe2ad10",
    },
    {
        "id": "TC-13",
        "type": "Functional",
        "req": "Edge Cases (AI-missed)",
        "objective": "Edge case: Verify physical shock resistance after light drop (handheld use scenario)",
        "precondition": "Fully charged; power OFF; hard floor (tile/concrete); open drop area; cold mode Toggle OFF",
        "steps": "1. Inspect housing, LED, grille, controls.\n2. Hold fan at 0.5 m above hard floor.\n3. Release fan once onto hard surface.\n4. Inspect for cracks, loose parts, rattling.\n5. Slide ON; Press speed +/−; Toggle cold mode; verify motor and LED.",
        "expected": "No structural crack; no loose parts; device runs normally after drop; cosmetic scuffs acceptable.",
        "actual": "Minor cosmetic scuff on housing corner; no cracks or loose parts; all functions normal.",
        "status": "Passed",
        "bug": None,
        "executed": False,
        "video": None,
        "ai": "No",
        "remark": "Student-added edge case; AI did not generate",
    },
    {
        "id": "TC-14",
        "type": "Functional",
        "req": "Battery & Runtime",
        "objective": "Verify low-battery warning or graceful shutdown before sudden stop",
        "precondition": "Battery ~15%; speed at level 70; cold mode Toggle OFF",
        "steps": "1. Slide power ON; Press speed + until LED shows 70.\n2. Run until device slows, warns, or shuts off.\n3. Record behavior in final 5 min.",
        "expected": "Visible low-battery indication before shutdown; no abrupt stop; cannot restart until charged.",
        "actual": "Device shut off suddenly at ~8% with no prior warning; could not restart until recharged.",
        "status": "Failed",
        "bug": "B005",
        "executed": False,
        "video": None,
        "ai": "Yes",
        "remark": "Defect logged — GitHub commit 09ecbc7",
    },
    {
        "id": "TC-15",
        "type": "Functional",
        "req": "Power & Charging",
        "objective": "Verify safe behavior when charging is interrupted mid-cycle",
        "precondition": "Battery ~40%; power slide switch at OFF; Type-C charger connected",
        "steps": "1. Begin charging; wait until ~60% (or 30 min).\n2. Unplug cable abruptly.\n3. Slide power ON; Press speed + until LED shows 50; run 5 min.\n4. Slide power OFF; reconnect charger.",
        "expected": "No damage or excess heat; device runs normally; charging resumes on reconnect.",
        "actual": "No swelling or excess heat; device ran normally at level 50; charging indicator resumed on reconnect.",
        "status": "Passed",
        "bug": None,
        "executed": False,
        "video": None,
        "ai": "Yes",
        "remark": None,
    },
]

PROMPTS = [
    (STUDENT_ID, "Cursor Auto Agent + Claude", "prompt-log.md / Claude share 4c7fb91c", "INCOMPLETE",
     "AI missed charge-while-use, battery indicator, drop test edge cases; wrong control interface initially.",
     "Replaced edge cases TC-11/12/13; corrected Slide/Press/Toggle controls; executed on physical device."),
    (STUDENT_ID, "Cursor Auto Agent + Claude", "istqb-mindmap-ai-generated.md / Claude share 78514e41", "INVALID",
     "ISTQB CTFL errors: audit as static review, functional suitability under non-functional, sandwich integration.",
     "Documented 3 corrections in hw1.md CLO G9.1."),
]


def clear_rows(ws, start_row, end_row, max_col=24):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).value = None


def update_requirement_list(ws):
    clear_rows(ws, 2, 6)
    for i, (no, rid, name, details) in enumerate(REQUIREMENTS):
        row = 2 + i
        ws.cell(row=row, column=1, value=no)
        ws.cell(row=row, column=2, value=rid)
        ws.cell(row=row, column=3, value=name)
        ws.cell(row=row, column=4, value=details)
        ws.cell(row=row, column=6, value=f"=COUNTIF('Test Cases'!D:D,C{row})")
    total_row = 2 + len(REQUIREMENTS)
    ws.cell(row=total_row, column=4, value="Total")
    ws.cell(row=total_row, column=5, value="Total")
    ws.cell(row=total_row, column=6, value=f"=SUM(F2:F{total_row - 1})")
    ws.cell(row=total_row + 1, column=1, value="<<< Insert New Line above this line >>>")


def update_test_types(ws):
    clear_rows(ws, 2, 7)
    for i, (no, tid, details) in enumerate(TEST_TYPES):
        row = 2 + i
        ws.cell(row=row, column=1, value=no)
        ws.cell(row=row, column=2, value=tid)
        ws.cell(row=row, column=3, value=details)
        ws.cell(row=row, column=5, value=f"=COUNTIF('Test Cases'!C:C,B{row})")
    total_row = 2 + len(TEST_TYPES)
    ws.cell(row=total_row, column=4, value="Total")
    ws.cell(row=total_row, column=5, value=f"=SUM(E2:E{total_row - 1})")
    ws.cell(row=total_row + 1, column=1, value="<<< Insert New Line above this line >>>")


def update_test_cases(ws):
    ws["M1"] = "Physical Test Execution — GOOJODOQ GFS006"
    ws["S1"] = "Physical Test Execution — GOOJODOQ GFS006 (copy)"
    clear_rows(ws, 4, 20)
    wrap = Alignment(wrap_text=True, vertical="top")
    for i, tc in enumerate(TEST_CASES):
        row = 4 + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=tc["id"])
        ws.cell(row=row, column=3, value=tc["type"])
        ws.cell(row=row, column=4, value=tc["req"])
        ws.cell(row=row, column=5, value=BUILD)
        ws.cell(row=row, column=6, value=tc["objective"])
        ws.cell(row=row, column=7, value=tc["precondition"])
        ws.cell(row=row, column=8, value=tc["steps"])
        ws.cell(row=row, column=10, value=tc["expected"])
        ws.cell(row=row, column=11, value=STUDENT_ID)
        ws.cell(row=row, column=12, value=tc["ai"])
        # Test Execution 2 (columns M-R) — matches Test Summary Report formulas (Status = col N)
        ws.cell(row=row, column=13, value=tc["actual"])
        ws.cell(row=row, column=14, value=tc["status"])
        ws.cell(row=row, column=15, value=tc["bug"])
        ws.cell(row=row, column=16, value=STUDENT_NAME if tc["executed"] or tc["actual"] != "As expected" else None)
        ws.cell(row=row, column=17, value=TEST_DATE if tc["executed"] or tc["status"] in ("Passed", "Failed") else None)
        remark_parts = []
        if tc.get("remark"):
            remark_parts.append(tc["remark"])
        if tc.get("video"):
            remark_parts.append(f"Video: {tc['video']}")
        if tc["executed"]:
            remark_parts.append("Executed with video: Yes")
        remark = "\n".join(remark_parts) if remark_parts else None
        ws.cell(row=row, column=18, value=remark)
        # Mirror to Test Execution 1 (columns S-X) for consistency
        ws.cell(row=row, column=19, value=tc["actual"])
        ws.cell(row=row, column=20, value=tc["status"])
        ws.cell(row=row, column=21, value=tc["bug"])
        ws.cell(row=row, column=22, value=ws.cell(row=row, column=16).value)
        ws.cell(row=row, column=23, value=ws.cell(row=row, column=17).value)
        ws.cell(row=row, column=24, value=remark)
        for col in (6, 7, 8, 10, 13, 18):
            ws.cell(row=row, column=col).alignment = wrap
    insert_row = 4 + len(TEST_CASES)
    ws.cell(row=insert_row, column=1, value="<<< Insert New Line above this line >>>")


def update_prompt_history(ws):
    clear_rows(ws, 2, 5)
    for i, row_data in enumerate(PROMPTS):
        row = 2 + i
        for col, val in enumerate(row_data, start=1):
            ws.cell(row=row, column=col, value=val)


def update_summary(ws):
    ws["C2"] = "HW01 - GOOJODOQ GFS006 Portable Fan Physical Testing"
    ws["C3"] = f"{STUDENT_NAME} ({STUDENT_ID})"
    ws["H2"] = "Dr. Lam Quang Vu / FIT@HCMUS"
    ws["H3"] = STUDENT_NAME
    ws["K6"] = TEST_DATE
    # Extend summary rows for 8 requirements
    req_count = len(REQUIREMENTS)
    # Clear old template rows 9-11 and rebuild
    for r in range(9, 20):
        for c in range(1, 12):
            ws.cell(row=r, column=c).value = None
    for i in range(req_count):
        row = 9 + i
        src = 2 + i
        ws.cell(row=row, column=1, value=i + 1)
        ws.cell(row=row, column=2, value=f"='Requirement List'!B{src}")
        ws.cell(row=row, column=3, value=f"='Requirement List'!C{src}")
        ws.cell(row=row, column=4, value=f"=SUM('Test Summary Report'!$E{row}:$G{row})")
        ws.cell(row=row, column=5, value=f"=COUNTIFS('Test Cases'!$D:$D,'Test Summary Report'!$C{row},'Test Cases'!$N:$N,'Test Summary Report'!E$8)")
        ws.cell(row=row, column=6, value=f"=COUNTIFS('Test Cases'!$D:$D,'Test Summary Report'!$C{row},'Test Cases'!$N:$N,'Test Summary Report'!F$8)")
        ws.cell(row=row, column=7, value=f"=COUNTIFS('Test Cases'!$D:$D,'Test Summary Report'!$C{row},'Test Cases'!$N:$N,'Test Summary Report'!G$8)")
        ws.cell(row=row, column=8, value=f"=COUNTIFS('Test Cases'!$D:$D,'Test Summary Report'!$C{row},'Test Cases'!$N:$N,'Test Summary Report'!H$8)")
        ws.cell(row=row, column=9, value=f"=COUNTIFS('Test Cases'!$D:$D,'Test Summary Report'!$C{row})-SUM(D{row},H{row})")
        ws.cell(row=row, column=10, value=f"=SUM('Test Summary Report'!$E{row}:$I{row})")
        ws.cell(row=row, column=11, value=f"='Test Summary Report'!$D{row}/'Test Summary Report'!$J{row}")
    total_row = 9 + req_count
    ws.cell(row=total_row, column=3, value="Total")
    ws.cell(row=total_row, column=4, value=f"=SUM(D9:D{total_row - 1})")
    ws.cell(row=total_row, column=5, value=f"=SUM(E9:E{total_row - 1})")
    ws.cell(row=total_row, column=6, value=f"=SUM(F9:F{total_row - 1})")
    ws.cell(row=total_row, column=7, value=f"=SUM(G9:G{total_row - 1})")
    ws.cell(row=total_row, column=8, value=f"=SUM(H9:H{total_row - 1})")
    ws.cell(row=total_row, column=9, value=f"=SUM(I9:I{total_row - 1})")
    ws.cell(row=total_row, column=10, value=f"=SUM(J9:J{total_row - 1})")
    ws.cell(row=total_row, column=11, value=f"=D{total_row}/J{total_row}")
    ws.cell(row=total_row + 1, column=1, value="<<< Insert New Line above this line >>>")
    # Update coverage formula references to new total row
    ws["C5"] = f"=D{total_row}/J{total_row}"
    ws["C6"] = f"=E{total_row}/J{total_row}"


def main():
    wb = openpyxl.load_workbook(PATH)
    update_requirement_list(wb["Requirement List"])
    update_test_types(wb["Test Types"])
    update_test_cases(wb["Test Cases"])
    update_prompt_history(wb["Prompt History"])
    update_summary(wb["Test Summary Report"])
    wb.save(PATH)
    print(f"Updated {PATH} with {len(TEST_CASES)} test cases.")


if __name__ == "__main__":
    main()
